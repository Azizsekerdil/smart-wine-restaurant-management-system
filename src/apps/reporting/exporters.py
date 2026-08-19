"""Rapor dışa aktarıcıları: PDF, Excel ve CSV.

TÜRKÇE KARAKTER GARANTİSİ
-------------------------
Türkçe'ye özgü karakterler (ç ğ ı İ ö ş ü) her üç biçimde de doğru
görünmelidir. Bunun için:

  * **CSV**   — UTF-8 BOM ile yazılır (``utf-8-sig``); Excel dosyayı
    çift tıklamayla açtığında karakterler bozulmaz.
  * **Excel** — openpyxl zaten UTF-8 çalışır; hücre biçimleri Türkçe
    tarih/para düzenine göre ayarlanır.
  * **PDF**   — ReportLab'ın gömülü Type-1 yazı tipleri (Helvetica) Latin-1
    ile sınırlıdır ve Türkçe'yi tam karşılamaz. Bu nedenle sistemde bulunan
    bir TrueType yazı tipi (DejaVu Sans veya Windows Arial/Tahoma) aranır
    ve gömülür. Uygun yazı tipi bulunamazsa dışa aktarım *sessizce bozuk
    çıktı üretmek yerine* açık bir hata verir.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

#: PDF için aranan TrueType yazı tipleri (öncelik sırasıyla)
FONT_CANDIDATES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("DejaVuSans", ("DejaVuSans.ttf",)),
    ("Arial", ("arial.ttf", "Arial.ttf")),
    ("Tahoma", ("tahoma.ttf", "Tahoma.ttf")),
    ("Verdana", ("verdana.ttf", "Verdana.ttf")),
    ("SegoeUI", ("segoeui.ttf",)),
)

#: Depoyla birlikte gelen yazı tipi klasörü.
#: DejaVu Sans (Bitstream Vera / Arev, izin verici lisans) depoda paketlidir;
#: böylece PDF çıktısı işletim sistemindeki tescilli bir yazı tipine (Arial,
#: Segoe UI) bağlı olmaz ve her kurulumda aynı görünür.
#: Lisans metni: src/static/fonts/LICENSE_DEJAVU.txt
BUNDLED_FONT_DIR = Path(__file__).resolve().parents[2] / "static" / "fonts"

FONT_SEARCH_DIRS: tuple[Path, ...] = (
    BUNDLED_FONT_DIR,
    Path("/usr/share/fonts/truetype/dejavu"),
    Path("/usr/share/fonts/truetype"),
    Path("C:/Windows/Fonts"),
    Path("/Library/Fonts"),
)


class ExportError(RuntimeError):
    """Dışa aktarım hatası."""


@dataclass
class ReportTable:
    """Dışa aktarılacak rapor verisi.

    Attributes:
        title: Rapor başlığı.
        columns: Sütun başlıkları.
        rows: Satırlar (her satır ``columns`` uzunluğunda olmalıdır).
        subtitle: Alt başlık (dönem bilgisi vb.).
        totals: Toplam satırı (isteğe bağlı).
        metadata: Rapor üstverisi (üreten kullanıcı, tarih vb.).
    """

    title: str
    columns: list[str]
    rows: list[list[Any]]
    subtitle: str = ""
    totals: list[Any] | None = None
    metadata: dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        width = len(self.columns)
        for index, row in enumerate(self.rows):
            if len(row) != width:
                raise ExportError(
                    f"Satır {index} sütun sayısıyla uyuşmuyor " f"({len(row)} ≠ {width})."
                )


# ---------------------------------------------------------------------------
# Biçimlendirme yardımcıları
# ---------------------------------------------------------------------------
def format_value(value: Any, *, language: str = "tr") -> str:
    """Değeri kullanıcı diline göre biçimlendirir."""
    if value is None:
        return ""
    if isinstance(value, bool):
        if language == "tr":
            return "Evet" if value else "Hayır"
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M" if language == "tr" else "%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y" if language == "tr" else "%Y-%m-%d")
    if isinstance(value, Decimal):
        return format_decimal(value, language=language)
    if isinstance(value, float):
        return format_decimal(Decimal(str(value)), language=language)
    return str(value)


def format_decimal(value: Decimal, *, language: str = "tr", places: int = 2) -> str:
    """Sayıyı yerel biçimde döndürür.

    Türkçe: ``1.234,56`` — binlik ayırıcı nokta, ondalık virgül.
    İngilizce: ``1,234.56``
    """
    # Mali gösterimde yarıyı yukarı yuvarlama beklenir (banker yuvarlaması değil).
    quantized = (
        value.quantize(Decimal("1." + "0" * places), rounding=ROUND_HALF_UP) if places else value
    )
    formatted = f"{quantized:,.{places}f}"
    if language == "tr":
        # İngilizce biçimden Türkçe biçime çevir (geçici işaretleyici ile)
        formatted = formatted.replace(",", "\x00").replace(".", ",").replace("\x00", ".")
    return formatted


def format_money(value: Decimal, *, language: str = "tr", symbol: str = "₺") -> str:
    """Para birimini yerel biçimde döndürür."""
    amount = format_decimal(Decimal(str(value)), language=language)
    return f"{amount} {symbol}" if language == "tr" else f"{symbol}{amount}"


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def export_csv(table: ReportTable, *, language: str = "tr", delimiter: str = ";") -> bytes:
    """CSV üretir.

    UTF-8 BOM (``utf-8-sig``) kullanılır; böylece Excel dosyayı çift
    tıklamayla açtığında Türkçe karakterler bozulmaz. Türkçe yerelde ondalık
    ayırıcı virgül olduğu için alan ayırıcı varsayılan olarak noktalı
    virgüldür.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)

    writer.writerow([table.title])
    if table.subtitle:
        writer.writerow([table.subtitle])
    writer.writerow([])
    writer.writerow(table.columns)

    for row in table.rows:
        writer.writerow([format_value(cell, language=language) for cell in row])

    if table.totals:
        writer.writerow([format_value(cell, language=language) for cell in table.totals])

    return buffer.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def export_excel(table: ReportTable, *, language: str = "tr") -> bytes:
    """XLSX üretir."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (table.title[:28] or "Rapor").replace("/", "-").replace("\\", "-")

    header_fill = PatternFill("solid", fgColor="4A1C2E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    total_font = Font(bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.cell(row=1, column=1, value=table.title).font = title_font
    current_row = 2
    if table.subtitle:
        sheet.cell(row=current_row, column=1, value=table.subtitle)
        current_row += 1
    current_row += 1

    header_row = current_row
    for index, column in enumerate(table.columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    current_row += 1

    number_format = "#,##0.00"
    date_format = "DD.MM.YYYY" if language == "tr" else "YYYY-MM-DD"

    for row in table.rows:
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=current_row, column=index)
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = number_format
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.value = value
                cell.number_format = number_format if isinstance(value, float) else "#,##0"
            elif isinstance(value, datetime):
                cell.value = value.replace(tzinfo=None)
                cell.number_format = f"{date_format} HH:MM"
            elif isinstance(value, date):
                cell.value = value
                cell.number_format = date_format
            else:
                cell.value = format_value(value, language=language)
            cell.border = border
        current_row += 1

    if table.totals:
        for index, value in enumerate(table.totals, start=1):
            cell = sheet.cell(row=current_row, column=index)
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = number_format
            else:
                cell.value = format_value(value, language=language)
            cell.font = total_font
            cell.border = border
        current_row += 1

    # Sütun genişlikleri
    for index, column in enumerate(table.columns, start=1):
        longest = len(str(column))
        for row in table.rows[:500]:
            longest = max(longest, len(format_value(row[index - 1], language=language)))
        sheet.column_dimensions[get_column_letter(index)].width = min(50, max(12, longest + 3))

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _find_unicode_font() -> tuple[str, Path]:
    """Türkçe karakterleri karşılayan bir TrueType yazı tipi bulur.

    Raises:
        ExportError: Uygun yazı tipi bulunamazsa.
    """
    for font_name, file_names in FONT_CANDIDATES:
        for directory in FONT_SEARCH_DIRS:
            if not directory.exists():
                continue
            for file_name in file_names:
                candidate = directory / file_name
                if candidate.exists():
                    return font_name, candidate

    raise ExportError(
        "PDF dışa aktarımı için Türkçe karakterleri destekleyen bir TrueType "
        "yazı tipi bulunamadı. Aranan konumlar: "
        + ", ".join(str(path) for path in FONT_SEARCH_DIRS)
        + ". Çözüm: DejaVuSans.ttf dosyasını sisteme kurun veya Windows'ta "
        "Arial/Tahoma yazı tipinin bulunduğundan emin olun."
    )


def _register_pdf_font() -> str:
    """Yazı tipini ReportLab'a kaydeder ve adını döndürür."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name, font_path = _find_unicode_font()
    registered = f"WH-{font_name}"
    if registered not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(registered, str(font_path)))
    return registered


def export_pdf(
    table: ReportTable,
    *,
    language: str = "tr",
    landscape_mode: bool = False,
    app_name: str = "Wine House",
) -> bytes:
    """PDF üretir.

    Türkçe karakterlerin doğru görünmesi için Unicode destekli bir TrueType
    yazı tipi gömülür.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font_name = _register_pdf_font()
    page_size = landscape(A4) if landscape_mode else A4

    stream = io.BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=18 * mm,
        title=table.title,
        author=app_name,
    )

    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "WHTitle",
        parent=base["Title"],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#4A1C2E"),
    )
    subtitle_style = ParagraphStyle(
        "WHSubtitle",
        parent=base["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#666666"),
    )
    cell_style = ParagraphStyle(
        "WHCell", parent=base["Normal"], fontName=font_name, fontSize=8, leading=10
    )
    header_style = ParagraphStyle(
        "WHHeader",
        parent=base["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    story: list[Any] = [Paragraph(table.title, title_style)]
    if table.subtitle:
        story.append(Paragraph(table.subtitle, subtitle_style))
    if table.metadata:
        meta_line = " · ".join(f"{key}: {value}" for key, value in table.metadata.items())
        story.append(Paragraph(meta_line, subtitle_style))
    story.append(Spacer(1, 6 * mm))

    data: list[list[Any]] = [[Paragraph(str(column), header_style) for column in table.columns]]
    for row in table.rows:
        data.append([Paragraph(format_value(cell, language=language), cell_style) for cell in row])
    if table.totals:
        data.append(
            [
                Paragraph(f"<b>{format_value(cell, language=language)}</b>", cell_style)
                for cell in table.totals
            ]
        )

    available_width = page_size[0] - document.leftMargin - document.rightMargin
    column_width = available_width / max(1, len(table.columns))

    pdf_table = Table(data, colWidths=[column_width] * len(table.columns), repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A1C2E")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F3F4")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if table.totals:
        style_commands.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EDE3E6")))
    pdf_table.setStyle(TableStyle(style_commands))
    story.append(pdf_table)

    def _footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        stamp = datetime.now().strftime("%d.%m.%Y %H:%M")
        canvas.drawString(doc.leftMargin, 10 * mm, f"{app_name} · {stamp}")
        canvas.drawRightString(page_size[0] - doc.rightMargin, 10 * mm, f"Sayfa {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return stream.getvalue()


# ---------------------------------------------------------------------------
# Ortak giriş noktası
# ---------------------------------------------------------------------------
CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def export(table: ReportTable, output_format: str, *, language: str = "tr") -> tuple[bytes, str]:
    """Raporu istenen biçimde üretir.

    Returns:
        ``(içerik, content_type)``

    Raises:
        ExportError: Desteklenmeyen biçim istenirse.
    """
    normalized = output_format.lower().strip()
    if normalized == "csv":
        return export_csv(table, language=language), CONTENT_TYPES["csv"]
    if normalized in {"xlsx", "excel"}:
        return export_excel(table, language=language), CONTENT_TYPES["xlsx"]
    if normalized == "pdf":
        return (
            export_pdf(table, language=language, landscape_mode=len(table.columns) > 6),
            CONTENT_TYPES["pdf"],
        )
    raise ExportError(f"Desteklenmeyen dışa aktarım biçimi: {output_format}")


def safe_filename(title: str, output_format: str) -> str:
    """Türkçe karakterleri ASCII'ye çeviren güvenli dosya adı üretir."""
    import re
    import unicodedata

    translation = str.maketrans("çÇğĞıİöÖşŞüÜ", "cCgGiIoOsSuU")
    ascii_title = title.translate(translation)
    ascii_title = unicodedata.normalize("NFKD", ascii_title).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^A-Za-z0-9]+", "-", ascii_title).strip("-").lower() or "rapor"
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return f"{slug}-{stamp}.{output_format.lower()}"
